"""
NOV 2-coat IR report generator.
Fills template_NOV_2coat_IR.xlsx with inspection data from the app's database.

Usage in FastAPI:
    from fill_nov_report import fill_report
    xlsx_path = fill_report(data, out_dir="/tmp")

Payload schema (all values strings unless noted; use "NA" for unused coats):
{
  "nov_po": "131923",
  "equipment_description": "...",
  "part_number": "10309693-001",
  "job_number": "50333868",
  "serial_no": "NA",
  "serial_no_end": "",             # optional; fills the "to <serial>" range
  "report_date": "15/11/2025",

  "surface_prep": {
    "method": "SSPC SP-10 Near White Blast Cleaning",
    "booth_temp": "79-80", "surface_temp": "79-80",
    "nozzle_pressure": "93", "dew_point": "71-72",
    "relative_humidity": "66-70%", "abrasive": "Aluminum Oxide",
    "blast_profile": "2.0-2.5", "time_started": "18:10",
    "accept": "OK", "date": "14/11/2025", "operator": "Mr. Zuber"
  },

  "batch_numbers": {"primer": "A400024958 / B400024951",
                    "second": "NA", "third": "NA", "fourth": "NA"},

  # exactly 4 rows: primer, 2nd, 3rd, 4th
  "applications": [
    {"product": "Carbozinc 858", "color": "Grey", "date": "14/11/2025",
     "time": "19:15:00", "operator": "Mr. Sudarshan B"},
    {"product": "NA", "color": "NA", "date": "NA", "time": "NA", "operator": "NA"},
    {"product": "NA", "color": "NA", "date": "NA", "time": "NA", "operator": "NA"},
    {"product": "NA", "color": "NA", "date": "NA", "time": "NA", "operator": "NA"}
  ],

  # exactly 4 rows, same coat order
  "conditions": [
    {"booth_temp": "79-80 °F", "surface_temp": "79-80 °F",
     "dew_point": "71-72 °F", "relative_humidity": "66-70%"},
    {"booth_temp": "NA", "surface_temp": "NA", "dew_point": "NA", "relative_humidity": "NA"},
    {"booth_temp": "NA", "surface_temp": "NA", "dew_point": "NA", "relative_humidity": "NA"},
    {"booth_temp": "NA", "surface_temp": "NA", "dew_point": "NA", "relative_humidity": "NA"}
  ],

  "cure_test": {"batch_no": "I251511", "accept": "OK",
                "date": "15/11/2025", "approver": "Mr. Pavan P"},
  "visual_inspection": {"accept": "OK", "date": "15/11/2025",
                        "approver": "Mr. Pavan P"},

  "nov_paint_spec": "MVG040014, SYSTEM 3.1,White/Gray (RAL 9002)",
  "dft": {
    # spec min/max per coat (strings; "NA" where unused)
    "spec":     {"primer": ["3", "5"], "second": ["NA", "NA"],
                 "third": ["2", "5"], "fourth": ["NA", "NA"]},
    "measured": {"primer": ["3.50", "5.00"], "second": ["NA", "NA"],
                 "third": ["", "3.30"], "fourth": ["NA", "NA"]},
    "total_min": "3.50", "total_max": "5.00"
  },

  # up to 4 gauges
  "gauges": [
    {"name": "PosiTector 6000 FNS Probe 1076838", "cal_due": "21/11/2025"},
    ...
  ],

  "completed_by": "Pavan Patil", "completed_date": "15/11/2025",
  "approved_by": "Ladislav Kiss", "approved_date": "15/11/2025"
}
"""
import os
import openpyxl

TEMPLATE = os.path.join(os.path.dirname(__file__), "..","report_templates", "NOV","template_NOV_2coat_IR.xlsx")

COAT_KEYS = ["primer", "second", "third", "fourth"]
APP_ROWS = [34, 35, 36, 37]        # applications table rows
COND_ROWS = [41, 42, 43, 44]       # conditions table rows
GAUGE_ROWS = [72, 73, 74, 75]
DFT_COLS = {"primer": ("E", "F"), "second": ("G", "H"),
            "third": ("I", "J"), "fourth": ("K", "L")}


def fill_report(data: dict, out_dir: str = ".", filename: str | None = None) -> str:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb["NOV 2 coat IR"]

    # ---- header / identification ----
    ws["J9"] = data["nov_po"]
    ws["C8"] = data["equipment_description"]
    ws["C11"] = data["part_number"]
    ws["C12"] = data["job_number"]
    ws["J11"] = data["serial_no"]
    end = data.get("serial_no_end", "")
    if end:
        ws["L11"] = "to"
        ws["J12"] = end
    ws["J8"] = data["report_date"]

    # ---- surface preparation ----
    sp = data["surface_prep"]
    ws["E15"] = sp["method"]
    ws["E16"] = sp["booth_temp"]
    ws["E17"] = sp["surface_temp"]
    ws["E18"] = sp["nozzle_pressure"]
    ws["E19"] = sp["dew_point"]
    ws["E20"] = sp["relative_humidity"]
    ws["E21"] = sp["abrasive"]
    ws["E22"] = sp["blast_profile"]
    ws["E23"] = sp["time_started"]
    ws["A26"] = sp["accept"]
    ws["E26"] = sp["date"]
    ws["I26"] = sp["operator"]

    # ---- batch numbers ----
    bn = data["batch_numbers"]
    ws["B29"] = bn["primer"]
    ws["B30"] = bn["second"]
    ws["H29"] = bn["third"]
    ws["H30"] = bn["fourth"]

    # ---- applications ----
    for row, app in zip(APP_ROWS, data["applications"]):
        ws[f"B{row}"] = app["product"]
        ws[f"D{row}"] = app["color"]
        ws[f"F{row}"] = app["date"]
        ws[f"H{row}"] = app["time"]
        ws[f"J{row}"] = app["operator"]

    # ---- conditions at start time ----
    for row, c in zip(COND_ROWS, data["conditions"]):
        ws[f"B{row}"] = c["booth_temp"]
        ws[f"D{row}"] = c["surface_temp"]
        ws[f"G{row}"] = c["dew_point"]
        ws[f"J{row}"] = c["relative_humidity"]

    # ---- cure test & visual inspection ----
    ct = data["cure_test"]
    ws["D50"] = ct["batch_no"]
    ws["A52"] = ct["accept"]
    ws["E52"] = ct["date"]
    ws["I52"] = ct["approver"]

    vi = data["visual_inspection"]
    ws["A57"] = vi["accept"]
    ws["E57"] = vi["date"]
    ws["I57"] = vi["approver"]

    # ---- DFT ----
    ws["E62"] = data["nov_paint_spec"]
    dft = data["dft"]
    for coat in COAT_KEYS:
        cmin, cmax = DFT_COLS[coat]
        ws[f"{cmin}67"] = dft["spec"][coat][0]
        ws[f"{cmax}67"] = dft["spec"][coat][1]
        ws[f"{cmin}68"] = dft["measured"][coat][0]
        ws[f"{cmax}68"] = dft["measured"][coat][1]
    ws["E69"] = dft["total_min"]
    ws["G69"] = dft["total_max"]

    # ---- gauges ----
    for row, g in zip(GAUGE_ROWS, data.get("gauges", [])):
        ws[f"B{row}"] = g["name"]
        ws[f"J{row}"] = g["cal_due"]

    # ---- signatures ----
    ws["B78"] = data["completed_by"]
    ws["J78"] = data["completed_date"]
    ws["B82"] = data["approved_by"]
    ws["J82"] = data["approved_date"]

    fname = filename or f"NOV_IR_{data['nov_po']}_{data['job_number']}.xlsx"
    out_path = os.path.join(out_dir, fname)
    wb.save(out_path)
    return out_path
