"""Final Paint Report generation (PDF + Excel).

Layout follows the NOV IR reference template (report_templates/nov_ir_format.csv):
header block (spec/system, document meta, company), equipment/part info,
surface preparation section, batch numbers, applications per coat, conditions
at start per coat, curing + visual results, DFT table, and sign-off rows.

Only the stages present on the work order appear — a WO's stage list already
reflects its case type, so per-case scoping needs no filtering here.

Measurements are reported in the units they were captured in (µm for DFT/WFT,
mils for surface/anchor profile); the reference template used mils throughout,
so column labels state the unit explicitly.
"""
import io
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

COAT_STAGES = ("primer_coat", "intermediate_coat", "top_coat")
COAT_LABELS = {"primer_coat": "Primer", "intermediate_coat": "Intermediate", "top_coat": "Top Coat"}
COAT_SUFFIX = {"primer_coat": "primer", "intermediate_coat": "intermediate", "top_coat": "top"}


def _stage(wo: dict, key: str) -> Optional[dict]:
    return next((s for s in wo["stages"] if s["key"] == key), None)


def _fmt_dt(iso: Optional[str]) -> str:
    if not iso:
        return "—"
    try:
        return datetime.fromisoformat(iso).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return iso


def _merged_fields(stage: dict) -> dict:
    sub = stage.get("submission") or {}
    return {**(stage.get("start_fields") or {}), **(sub.get("fields") or {})}


def _readings_str(r: Optional[dict]) -> str:
    if not r:
        return "—"
    return (f"Air {r.get('ambient_temp_c', '—')}°C · Surface {r.get('surface_temp_c', '—')}°C · "
            f"Dew {r.get('dew_point_c', '—')}°C · RH {r.get('relative_humidity_pct', '—')}%")


def collect_report_data(wo: dict, company: Optional[dict]) -> dict:
    """Normalize a work order (with stages) into the report section model."""
    sp = _stage(wo, "surface_prep")
    qa = _stage(wo, "curing_qa")
    qa_fields = _merged_fields(qa) if qa else {}
    coats = [s for k in COAT_STAGES if (s := _stage(wo, k))]

    sp_fields = _merged_fields(sp) if sp else {}
    sp_sub = (sp or {}).get("submission") or {}

    applications = []
    conditions = []
    batches = []
    dft_rows = []
    for s in coats:
        f = _merged_fields(s)
        label = COAT_LABELS[s["key"]]
        applications.append({
            "coat": label,
            "product": f.get("product", "—"),
            "brand": f.get("brand", "—"),
            "color": f.get("color") or f.get("paint_shade") or f.get("ral_shade") or "—",
            "date": _fmt_dt(s.get("submitted_at")),
            "start_time": f.get("process_start_time") or _fmt_dt(s.get("started_at")),
            "end_time": f.get("process_end_time") or "—",
            "operator": f.get("operator_name", "—"),
            "operator_designation": f.get("operator_designation", "—"),
            "visual": f.get("visual_inspection", "—"),
            "result": (s.get("result") or "pending").upper(),
        })
        conditions.append({"coat": label, "readings": _readings_str(s.get("start_readings"))})
        suffix = COAT_SUFFIX[s["key"]]
        # batch/expiry are captured at the coat stage since migration 0010;
        # older WOs carry them on curing_qa (legacy per-coat keys)
        batches.append({
            "coat": label,
            "batch": f.get("batch_number") or qa_fields.get(f"batch_number_{suffix}", "—"),
            "expiry": f.get("expiry_date") or qa_fields.get(f"expiry_date_{suffix}", "—"),
        })
        window = None
        if s.get("dft_window") and (wo.get("coat_limits") or {}).get(s["dft_window"]):
            window = wo["coat_limits"][s["dft_window"]]
        dft_rows.append({
            "coat": label,
            "min": window[0] if window else "—",
            "max": window[1] if window else "—",
            "measured": f.get("dft_um", "—"),
            "wft": f.get("wft_um", "—"),
        })

    failures = []
    for s in wo["stages"]:
        errs = ((s.get("submission") or {}).get("errors")) or []
        for e in errs:
            failures.append(f"{s['name']}: {e}")

    return {
        "title": f"FINAL PAINT REPORT — {wo['paint_product_code']}",
        "company": (company or {}).get("company_name") or "—",
        "logo_url": (company or {}).get("logo_url"),
        "meta": [
            ("Work Order", wo["work_order_id"]),
            ("Case Type", wo["case_type"].replace("_", " ").title()),
            ("Customer", wo["customer_name"]),
            ("Customer PO", f"{wo['po_number']} · line {wo.get('po_line_item_number') or '—'}"),
            ("Part", wo["part_description"]),
            ("Quantity / Serials", f"{wo['quantity']} · {wo['serial_range']}"),
            ("Specification", f"{wo['paint_product_code']} Rev {wo.get('coating_spec_revision_number') or '—'}"),
            ("Paint System", wo["paint_product_name"]),
            ("Report Date", datetime.now().strftime("%d/%m/%Y")),
        ],
        "surface_prep": {
            "present": sp is not None,
            "result": ((sp or {}).get("result") or "pending").upper(),
            "oil_water": sp_fields.get("oil_water_test", "—"),
            "profile": f"{sp_fields.get('surface_profile_mils', '—')} mils (anchor {sp_fields.get('anchor_profile_mils', '—')} mils)",
            "readings_start": _readings_str((sp or {}).get("start_readings")),
            "readings_end": _readings_str(((sp_sub.get("readings") or {}).get("end"))),
            "started": _fmt_dt((sp or {}).get("started_at")),
            "submitted": _fmt_dt((sp or {}).get("submitted_at")),
            "by": (sp or {}).get("submitted_by") or "—",
        },
        "batches": batches,
        "applications": applications,
        "conditions": conditions,
        "dft_rows": dft_rows,
        "qa": {
            "present": qa is not None,
            "mek": qa_fields.get("mek_test", "—"),
            "curing_room_temp": qa_fields.get("curing_room_temp", "—"),
            "adhesion": qa_fields.get("adhesion_tape", "—"),
            "result": ((qa or {}).get("result") or "pending").upper(),
            "by": (qa or {}).get("submitted_by") or "—",
            "date": _fmt_dt((qa or {}).get("submitted_at")),
        },
        "photos": {
            s["name"]: {
                "start": len(s.get("start_photos") or []),
                "end": len(((s.get("submission") or {}).get("photos")) or []),
            }
            for s in wo["stages"]
        },
        "failures": failures,
        "overall": "FAIL" if any(s.get("status") == "fail" for s in wo["stages"])
                   else ("PASS" if all(s.get("status") == "done" for s in wo["stages"]) else "IN PROGRESS"),
    }


# ---------------- Excel ----------------
def generate_xlsx(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Paint Report"
    bold = Font(bold=True)
    section = Font(bold=True, size=12)
    thin = Border(bottom=Side(style="thin", color="999999"))
    row = 1

    def put(values, fonts=None):
        nonlocal row
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            if fonts and col in fonts:
                c.font = fonts[col]
        row += 1

    def header(text):
        nonlocal row
        row += 1
        c = ws.cell(row=row, column=1, value=text)
        c.font = section
        c.border = thin
        row += 1

    ws.cell(row=row, column=1, value=data["title"]).font = Font(bold=True, size=14)
    ws.cell(row=row, column=6, value=data["company"]).font = bold
    row += 2
    for k, v in data["meta"]:
        put([k, None, v], fonts={1: bold})

    if data["surface_prep"]["present"]:
        sp = data["surface_prep"]
        header("SURFACE PREPARATION")
        put(["Result", None, sp["result"], None, "Oil/Water Test", sp["oil_water"]], fonts={1: bold, 5: bold})
        put(["Profile", None, sp["profile"]], fonts={1: bold})
        put(["Conditions (start)", None, sp["readings_start"]], fonts={1: bold})
        put(["Conditions (end)", None, sp["readings_end"]], fonts={1: bold})
        put(["Started", None, sp["started"], None, "Submitted", sp["submitted"], None, "By", sp["by"]],
            fonts={1: bold, 5: bold, 8: bold})

    if data["batches"]:
        header("BATCH NUMBERS")
        put(["Coat", "Batch Number", "Expiry Date"], fonts={1: bold, 2: bold, 3: bold})
        for b in data["batches"]:
            put([b["coat"], b["batch"], b["expiry"]])

    if data["applications"]:
        header("APPLICATIONS")
        put(["Coat", "Brand", "Product", "Color/Shade", "Date", "Start", "End", "Operator", "Designation", "Visual", "Result"],
            fonts={i: bold for i in range(1, 12)})
        for a in data["applications"]:
            put([a["coat"], a["brand"], a["product"], a["color"], a["date"], a["start_time"],
                 a["end_time"], a["operator"], a["operator_designation"], a["visual"], a["result"]])

        header("CONDITIONS AT START OF COAT")
        for c in data["conditions"]:
            put([c["coat"], None, c["readings"]], fonts={1: bold})

        header("DRY FILM THICKNESS (µm)")
        put(["Coat", "Min", "Max", "Measured DFT", "WFT"], fonts={i: bold for i in range(1, 6)})
        for d in data["dft_rows"]:
            put([d["coat"], d["min"], d["max"], d["measured"], d["wft"]])

    if data["qa"]["present"]:
        header("CURING + QA")
        qa = data["qa"]
        put(["MEK Resistance", qa["mek"], None, "Curing at Room Temp", qa["curing_room_temp"]], fonts={1: bold, 4: bold})
        put(["Adhesion Tape", qa["adhesion"]], fonts={1: bold})
        put(["Result", qa["result"], None, "By", qa["by"], None, "Date", qa["date"]], fonts={1: bold, 4: bold, 7: bold})

    header("EVIDENCE PHOTOS (count per stage)")
    for name, p in data["photos"].items():
        put([name, f"start: {p['start']}", f"end: {p['end']}"], fonts={1: bold})

    if data["failures"]:
        header("RECORDED FAILURE REASONS")
        for f in data["failures"]:
            put([f])

    header(f"OVERALL: {data['overall']}")
    ws.column_dimensions["A"].width = 26
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------- PDF ----------------
def generate_pdf(data: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    h1 = styles["Title"]
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], spaceBefore=10, spaceAfter=4)
    body = styles["BodyText"]

    def table(rows: List[List[str]], header_row=True, widths=None):
        t = Table(rows, colWidths=widths, hAlign="LEFT")
        style = [
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, rl_colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]
        if header_row:
            style += [("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#e8e8e8")),
                      ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")]
        t.setStyle(TableStyle(style))
        return t

    story = [Paragraph(data["title"], h1),
             Paragraph(f"Company: {data['company']}", body),
             Spacer(1, 4)]
    story.append(table([[k, str(v)] for k, v in data["meta"]], header_row=False, widths=[45 * mm, 130 * mm]))

    if data["surface_prep"]["present"]:
        sp = data["surface_prep"]
        story.append(Paragraph("Surface Preparation", h2))
        story.append(table([
            ["Result", sp["result"], "Oil/Water Test", sp["oil_water"]],
            ["Profile", sp["profile"], "By", sp["by"]],
            ["Conditions (start)", sp["readings_start"], "Started", sp["started"]],
            ["Conditions (end)", sp["readings_end"], "Submitted", sp["submitted"]],
        ], header_row=False))

    if data["batches"]:
        story.append(Paragraph("Batch Numbers", h2))
        story.append(table([["Coat", "Batch Number", "Expiry Date"]] +
                           [[b["coat"], b["batch"], b["expiry"]] for b in data["batches"]]))

    if data["applications"]:
        story.append(Paragraph("Applications", h2))
        story.append(table(
            [["Coat", "Product", "Color/Shade", "Date", "Start", "End", "Operator", "Visual", "Result"]] +
            [[a["coat"], f"{a['brand']} {a['product']}", a["color"], a["date"], a["start_time"],
              a["end_time"], f"{a['operator']} ({a['operator_designation']})", a["visual"], a["result"]]
             for a in data["applications"]]))
        story.append(Paragraph("Conditions at Start of Coat", h2))
        story.append(table([[c["coat"], c["readings"]] for c in data["conditions"]], header_row=False))
        story.append(Paragraph("Dry Film Thickness (µm)", h2))
        story.append(table([["Coat", "Min", "Max", "Measured DFT", "WFT"]] +
                           [[d["coat"], str(d["min"]), str(d["max"]), str(d["measured"]), str(d["wft"])]
                            for d in data["dft_rows"]]))

    if data["qa"]["present"]:
        qa = data["qa"]
        story.append(Paragraph("Curing + QA", h2))
        story.append(table([
            ["MEK Resistance", qa["mek"], "Curing at Room Temp", qa["curing_room_temp"]],
            ["Adhesion Tape", qa["adhesion"], "Result", qa["result"]],
            ["By", qa["by"], "Date", qa["date"]],
        ], header_row=False))

    story.append(Paragraph("Evidence Photos (count per stage)", h2))
    story.append(table([[name, f"start: {p['start']}", f"end: {p['end']}"]
                        for name, p in data["photos"].items()], header_row=False))

    if data["failures"]:
        story.append(Paragraph("Recorded Failure Reasons", h2))
        for f in data["failures"]:
            story.append(Paragraph(f"• {f}", body))

    story.append(Spacer(1, 8))
    story.append(Paragraph(f"OVERALL: {data['overall']}", h2))
    doc.build(story)
    return buf.getvalue()
