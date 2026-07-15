"""End-to-end backend tests for the Coating Portal API (case-type workflows).

Run against a live server + persistent DB: every test that needs a work order
creates its own (unique PO number per run), so the suite is rerunnable.
"""
import itertools
import os
import time
import pytest
import requests

# Port 8002 = the dedicated test stack provisioned by conftest.py (throwaway
# Docker Postgres + its own uvicorn). Never point this at the live backend.
BASE = os.environ.get("TEST_BACKEND_URL", "http://localhost:8002").rstrip("/")
API = f"{BASE}/api"

EMAIL = "j.thompson@aerospace-precision.com"
PASSWORD = "Inspector@123"

CASE_SEQUENCES = {
    "only_primer": ["surface_prep", "primer_coat", "curing_qa"],
    "primer_intermediate": ["surface_prep", "primer_coat", "intermediate_coat", "curing_qa"],
    "primer_intermediate_top": ["surface_prep", "primer_coat", "intermediate_coat", "top_coat", "curing_qa"],
    "top_coat_only": ["surface_prep", "top_coat", "curing_qa"],
}

_seq = itertools.count(1)


@pytest.fixture(scope="session")
def token():
    r = requests.post(f"{API}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()["access_token"]


@pytest.fixture
def auth(token):
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="session")
def spec(token):
    """A paint-system spec that has primer, intermediate, top and total DFT windows."""
    r = requests.get(f"{API}/coating-specifications",
                     headers={"Authorization": f"Bearer {token}"}, timeout=15)
    assert r.status_code == 200, r.text
    rows = r.json()
    return next(x for x in rows
                if x["specification"] == "MVG040014" and x["paint_brand"] == "JOTUN"
                and x["system_number"] == 2)


def _create_wo(auth, spec, case_type, **overrides):
    body = {
        "case_type": case_type,
        "customer_name": f"Test {case_type}",
        "po_number": f"PO-T-{int(time.time())}-{next(_seq)}",
        "po_line_item_number": 1,
        "part_number": f"PN-{case_type}",
        "part_revision_number": "A",
        "coating_spec_code": spec["specification"],
        "coating_spec_revision_number": spec["spec_rev"] or "0",
        "paint_system_id": spec["id"],
        "quantity": 1,
        **overrides,
    }
    r = requests.post(f"{API}/work-orders", json=body, headers=auth, timeout=15)
    assert r.status_code == 201, r.text
    return r.json()


def _detail(auth, wo_id):
    r = requests.get(f"{API}/work-orders/{wo_id}", headers=auth, timeout=15)
    assert r.status_code == 200, r.text
    return r.json()


def _readings(end_surface=18.0, end_dew=9.5):
    return {
        "start": {"ambient_temp_c": 22.0, "relative_humidity_pct": 45.0, "dew_point_c": 9.5, "surface_temp_c": 18.0},
        "end":   {"ambient_temp_c": 22.5, "relative_humidity_pct": 46.0, "dew_point_c": end_dew, "surface_temp_c": end_surface},
    }


def _submit(auth, wo_id, stage_key, fields, result="pass", readings=None):
    body = {"readings": readings or _readings(), "fields": fields,
            "notes": "TEST", "photos": [], "result": result}
    return requests.post(f"{API}/work-orders/{wo_id}/stages/{stage_key}/submit",
                         json=body, headers=auth, timeout=15)


def _stage_of(detail, stage_key):
    return next(s for s in detail["stages"] if s["key"] == stage_key)


def _valid_fields(detail, stage_key, phase=None, **overrides):
    """Build a passing field payload from the stage's own field definitions.

    phase=None -> all fields (single-shot submissions); "start"/"end" filter
    to the capture phase (two-step flow).
    """
    stage = _stage_of(detail, stage_key)
    vals = {}
    defs = [f for f in stage["fields"] if phase is None or f.get("phase", "end") == phase]
    for f in defs:
        t, k, opts = f["type"], f["key"], f.get("options", "")
        if t == "time":
            vals[k] = "08:30"
        elif t == "ok_notok":
            vals[k] = "OK"
        elif t == "pass_fail":
            vals[k] = "PASS"
        elif t == "date":
            vals[k] = "2027-01-01"
        elif t in ("text", "note"):
            vals[k] = "Test entry"
        elif t == "dropdown":
            if opts == "brands":
                vals[k] = "JOTUN"
            elif opts == "products.primer":
                vals[k] = "RESIST 86"
            elif opts == "products.intermediate":
                vals[k] = "JOTACOTE UNIVERSAL N10"
            elif opts == "products.top":
                vals[k] = "JOTAMASTIC 90"
            elif opts == "colors":
                vals[k] = "Gray"
            elif opts == "operators":
                vals[k] = "Mr.Kishore"       # seeded in migration 0007
            elif opts == "operator_designations":
                vals[k] = "Painter"
            # shades / ral are optional stubs with no options yet — skip
        elif t in ("number", "decimal"):
            rng = f.get("range")
            if rng == "pct":
                vals[k] = 68
            elif rng == "anchor_profile":
                lo = detail["spec"]["surface_profile_min_um"] / 25.4
                hi = detail["spec"]["surface_profile_max_um"] / 25.4
                vals[k] = round((lo + hi) / 2, 2)
            elif rng == "dft_window":
                lo, hi = detail["coat_limits"][stage["dft_window"]]
                vals[k] = (lo + hi) / 2
            elif rng == "wft":
                lo, hi = detail["coat_limits"][stage["dft_window"]]
                vals[k] = round(((lo + hi) / 2) * 100 / 68, 1)  # solids=68 above
            else:
                vals[k] = 1
    vals.update(overrides)
    return {k: v for k, v in vals.items() if v is not None}


# ---------- Auth ----------
class TestAuth:
    def test_login_success(self):
        r = requests.post(f"{API}/auth/login", json={"email": EMAIL, "password": PASSWORD}, timeout=15)
        assert r.status_code == 200
        body = r.json()
        assert body["token_type"] == "bearer"
        assert body["access_token"]
        assert body["user"]["employee_id"] == "QC-7742"
        assert body["user"]["role"] == "Lead Inspector"

    def test_login_wrong_password(self):
        r = requests.post(f"{API}/auth/login", json={"email": EMAIL, "password": "wrong"}, timeout=15)
        assert r.status_code == 401

    def test_me_requires_token(self):
        r = requests.get(f"{API}/auth/me", timeout=15)
        assert r.status_code == 401

    def test_me_with_token(self, auth):
        r = requests.get(f"{API}/auth/me", headers=auth, timeout=15)
        assert r.status_code == 200
        assert r.json()["employee_id"] == "QC-7742"


# ---------- Case types ----------
class TestCaseTypes:
    def test_four_case_types_with_exact_sequences(self, auth):
        r = requests.get(f"{API}/case-types", headers=auth, timeout=15)
        assert r.status_code == 200
        got = {ct["case_type"]: [s["key"] for s in ct["stages"]] for ct in r.json()}
        assert got == CASE_SEQUENCES

    def test_curing_qa_is_observational_in_all_cases(self, auth):
        r = requests.get(f"{API}/case-types", headers=auth, timeout=15)
        for ct in r.json():
            qa = next(s for s in ct["stages"] if s["key"] == "curing_qa")
            assert qa["params"] == []


# ---------- Work order creation ----------
class TestCreateWorkOrder:
    @pytest.mark.parametrize("case_type", list(CASE_SEQUENCES))
    def test_stages_generated_per_case_type(self, auth, spec, case_type):
        created = _create_wo(auth, spec, case_type)
        assert created["case_type"] == case_type
        assert created["total_stages"] == len(CASE_SEQUENCES[case_type])
        detail = _detail(auth, created["work_order_id"])
        assert [s["key"] for s in detail["stages"]] == CASE_SEQUENCES[case_type]
        assert all(s["status"] == "pending" for s in detail["stages"])

    def test_case_type_required(self, auth, spec):
        body = {
            "customer_name": "No Case", "po_number": f"PO-NC-{int(time.time())}",
            "po_line_item_number": 1, "part_number": "PN-NC", "part_revision_number": "A",
            "coating_spec_code": spec["specification"], "coating_spec_revision_number": "N",
            "paint_system_id": spec["id"], "quantity": 1,
        }
        r = requests.post(f"{API}/work-orders", json=body, headers=auth, timeout=15)
        assert r.status_code == 422  # pydantic: missing/invalid case_type

    def test_invalid_case_type_rejected(self, auth, spec):
        r = requests.post(f"{API}/work-orders", json={
            "case_type": "full_coating", "customer_name": "Bad Case",
            "po_number": f"PO-BC-{int(time.time())}", "po_line_item_number": 1,
            "part_number": "PN-BC", "part_revision_number": "A",
            "coating_spec_code": spec["specification"], "coating_spec_revision_number": "N",
            "paint_system_id": spec["id"], "quantity": 1,
        }, headers=auth, timeout=15)
        assert r.status_code == 422

    def test_duplicate_guard(self, auth, spec):
        po = f"PO-DUP-{int(time.time())}"
        _create_wo(auth, spec, "only_primer", po_number=po, part_number="PN-DUP")
        r = requests.post(f"{API}/work-orders", json={
            "case_type": "only_primer", "customer_name": "Dup Co", "po_number": po,
            "po_line_item_number": 1, "part_number": "PN-DUP", "part_revision_number": "A",
            "coating_spec_code": spec["specification"], "coating_spec_revision_number": "N",
            "paint_system_id": spec["id"], "quantity": 1,
        }, headers=auth, timeout=15)
        assert r.status_code == 409
        assert r.json()["detail"]["duplicate"] is True

    def test_detail_shape(self, auth, spec):
        created = _create_wo(auth, spec, "primer_intermediate_top")
        wo = _detail(auth, created["work_order_id"])
        assert wo["spec"]["dft_min_um"] > 0
        for window in ("primer", "intermediate", "top", "mid_cumulative", "total"):
            assert wo["coat_limits"][window], f"missing coat_limits.{window}"
        assert wo["po_number"].startswith("PO-T-")

    def test_detail_404(self, auth):
        r = requests.get(f"{API}/work-orders/WO-NOPE", headers=auth, timeout=15)
        assert r.status_code == 404


# ---------- Work order listing ----------
class TestWorkOrders:
    def test_list_contains_created_and_valid_format(self, auth, spec):
        created = _create_wo(auth, spec, "top_coat_only")
        r = requests.get(f"{API}/work-orders", headers=auth, timeout=15)
        assert r.status_code == 200
        items = r.json()
        ids = {w["work_order_id"] for w in items}
        assert created["work_order_id"] in ids
        import re
        for w in items:
            assert re.match(r"^WO-\d{4}-\d{4}$", w["work_order_id"]), w["work_order_id"]
            assert w["case_type"] in CASE_SEQUENCES
            assert w["total_stages"] == len(CASE_SEQUENCES[w["case_type"]])

    def test_search_q(self, auth, spec):
        created = _create_wo(auth, spec, "only_primer", customer_name="Searchable Unique Co")
        r = requests.get(f"{API}/work-orders?q=Searchable Unique", headers=auth, timeout=15)
        assert r.status_code == 200
        assert any(w["work_order_id"] == created["work_order_id"] for w in r.json())


# ---------- Stage validation per case (field definitions) ----------
class TestStageValidation:
    def test_surface_prep_profile_out_of_window_fails(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        # anchor window for this spec is 1-2.5 mils; 0.2 is below it
        r = _submit(auth, wo["work_order_id"], "surface_prep",
                    _valid_fields(detail, "surface_prep", surface_profile_mils=0.2))
        assert r.status_code == 200, r.text
        assert r.json()["result"] == "fail"

    def test_surface_prep_oil_water_not_ok_fails(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        r = _submit(auth, wo["work_order_id"], "surface_prep",
                    _valid_fields(detail, "surface_prep", oil_water_test="NOT_OK"))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"

    def test_missing_required_end_field_rejected(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        fields = _valid_fields(detail, "primer_coat")
        del fields["wft_um"]  # end-phase required
        r = _submit(auth, wo["work_order_id"], "primer_coat", fields)
        assert r.status_code == 400
        assert "wft_um" in r.json()["detail"]

    def test_missing_required_start_field_rejected_at_start(self, auth, spec):
        # operator/paint identification is enforced when the stage STARTS
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        start_fields = _valid_fields(detail, "primer_coat", phase="start")
        del start_fields["operator_name"]
        r = requests.post(
            f"{API}/work-orders/{wo['work_order_id']}/stages/primer_coat/start",
            json={"readings": _readings()["start"], "fields": start_fields},
            headers=auth, timeout=15)
        assert r.status_code == 400
        assert "operator_name" in r.json()["detail"]

    def test_primer_window_hard_block(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        _, primer_hi = detail["coat_limits"]["primer"]
        r = _submit(auth, wo["work_order_id"], "primer_coat",
                    _valid_fields(detail, "primer_coat", dft_um=primer_hi + 100))
        assert r.status_code == 422, r.text
        assert r.json()["detail"]["hard_block"] is True

    def test_primer_visual_not_ok_fails(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        r = _submit(auth, wo["work_order_id"], "primer_coat",
                    _valid_fields(detail, "primer_coat", visual_inspection="NOT_OK"))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"

    def test_wft_out_of_derived_window_fails(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        # WFT window = DFT window / (solids/100); 9999 is far above any window
        r = _submit(auth, wo["work_order_id"], "primer_coat",
                    _valid_fields(detail, "primer_coat", wft_um=9999))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"

    def test_intermediate_uses_cumulative_window(self, auth, spec):
        wo = _create_wo(auth, spec, "primer_intermediate")
        detail = _detail(auth, wo["work_order_id"])
        stage = _stage_of(detail, "intermediate_coat")
        assert stage["dft_window"] == "mid_cumulative"
        r = _submit(auth, wo["work_order_id"], "intermediate_coat",
                    _valid_fields(detail, "intermediate_coat"))
        assert r.status_code == 200, r.text
        assert r.json()["result"] == "pass"

    def test_top_coat_only_uses_standalone_top_window(self, auth, spec):
        wo = _create_wo(auth, spec, "top_coat_only")
        detail = _detail(auth, wo["work_order_id"])
        lo, hi = detail["coat_limits"]["top"]
        total_hi = detail["coat_limits"]["total"][1]
        assert hi < total_hi  # the standalone window is tighter than full-system
        assert _stage_of(detail, "top_coat")["dft_window"] == "top"
        ok = _submit(auth, wo["work_order_id"], "top_coat", _valid_fields(detail, "top_coat"))
        assert ok.status_code == 200 and ok.json()["result"] == "pass"
        # over the top window (but under total) must still hard-block
        wo2 = _create_wo(auth, spec, "top_coat_only")
        detail2 = _detail(auth, wo2["work_order_id"])
        blocked = _submit(auth, wo2["work_order_id"], "top_coat",
                          _valid_fields(detail2, "top_coat", dft_um=hi + 50))
        assert blocked.status_code == 422
        assert blocked.json()["detail"]["hard_block"] is True

    def test_curing_qa_requires_batch_expiry_at_start_per_case(self, auth, spec):
        # batch/expiry live on curing_qa's START phase (captured before work)
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        start_defs = [f for f in _stage_of(detail, "curing_qa")["fields"] if f.get("phase") == "start"]
        keys = {f["key"] for f in start_defs}
        assert {"batch_number_primer", "expiry_date_primer"} <= keys
        assert "batch_number_top" not in keys
        r = requests.post(
            f"{API}/work-orders/{wo['work_order_id']}/stages/curing_qa/start",
            json={"readings": _readings()["start"], "fields": {}}, headers=auth, timeout=15)
        assert r.status_code == 400
        assert "batch_number_primer" in r.json()["detail"]
        # top_coat_only: top batch/expiry required, no primer keys; full two-step passes
        wo2 = _create_wo(auth, spec, "top_coat_only")
        detail2 = _detail(auth, wo2["work_order_id"])
        keys2 = {f["key"] for f in _stage_of(detail2, "curing_qa")["fields"] if f.get("phase") == "start"}
        assert {"batch_number_top", "expiry_date_top"} <= keys2
        assert "batch_number_primer" not in keys2
        r2 = requests.post(
            f"{API}/work-orders/{wo2['work_order_id']}/stages/curing_qa/start",
            json={"readings": _readings()["start"],
                  "fields": _valid_fields(detail2, "curing_qa", phase="start")},
            headers=auth, timeout=15)
        assert r2.status_code == 200, r2.text
        ok = _submit(auth, wo2["work_order_id"], "curing_qa",
                     _valid_fields(detail2, "curing_qa", phase="end"))
        assert ok.status_code == 200, ok.text
        assert ok.json()["result"] == "pass"
        # merged record contains both phases
        stg = _stage_of(_detail(auth, wo2["work_order_id"]), "curing_qa")
        assert stg["submission"]["fields"].get("batch_number_top")
        assert stg["submission"]["fields"].get("mek_test") == "PASS"

    def test_curing_qa_mek_fail_marks_fail(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        r = _submit(auth, wo["work_order_id"], "curing_qa",
                    _valid_fields(detail, "curing_qa", mek_test="FAIL"))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"

    def test_curing_qa_bad_expiry_date_blocked_at_start(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        bad = _valid_fields(detail, "curing_qa", phase="start", expiry_date_primer="15-03-2027")
        r = requests.post(
            f"{API}/work-orders/{wo['work_order_id']}/stages/curing_qa/start",
            json={"readings": _readings()["start"], "fields": bad}, headers=auth, timeout=15)
        assert r.status_code == 422
        assert any("invalid date" in e for e in r.json()["detail"]["errors"])

    def test_stage_not_in_case_404(self, auth, spec):
        wo = _create_wo(auth, spec, "top_coat_only")  # has no primer_coat
        r = _submit(auth, wo["work_order_id"], "primer_coat", {})
        assert r.status_code == 404

    def test_gate_dew_point_fails_with_specific_reason(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        r = _submit(auth, wo["work_order_id"], "primer_coat", _valid_fields(detail, "primer_coat"),
                    readings=_readings(end_surface=10.0, end_dew=9.8))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"
        assert any("dew point" in e for e in r.json()["errors"])

    def test_gate_too_hot_fails_with_specific_reason(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        r = _submit(auth, wo["work_order_id"], "primer_coat", _valid_fields(detail, "primer_coat"),
                    readings=_readings(end_surface=65.0, end_dew=9.8))
        assert r.status_code == 200
        assert r.json()["result"] == "fail"
        assert any("Too hot for coat" in e for e in r.json()["errors"])

    def test_start_blocked_when_too_hot(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        r = requests.post(
            f"{API}/work-orders/{wo['work_order_id']}/stages/surface_prep/start",
            json={"readings": {"ambient_temp_c": 40, "relative_humidity_pct": 30,
                               "dew_point_c": 15, "surface_temp_c": 65}},
            headers=auth, timeout=15)
        assert r.status_code == 422
        assert any("Too hot for coat" in e for e in r.json()["detail"]["errors"])

    def test_start_blocked_when_too_close_to_dew_point(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        r = requests.post(
            f"{API}/work-orders/{wo['work_order_id']}/stages/surface_prep/start",
            json={"readings": {"ambient_temp_c": 20, "relative_humidity_pct": 80,
                               "dew_point_c": 14, "surface_temp_c": 16}},
            headers=auth, timeout=15)
        assert r.status_code == 422
        assert any("dew point" in e for e in r.json()["detail"]["errors"])


# ---------- Two-step start/end flow ----------
class TestTwoStepFlow:
    def test_start_then_submit(self, auth, spec):
        wo = _create_wo(auth, spec, "primer_intermediate")
        wo_id = wo["work_order_id"]
        detail = _detail(auth, wo_id)
        start_body = {
            "readings": {"ambient_temp_c": 21.5, "relative_humidity_pct": 44.0,
                         "dew_point_c": 9.0, "surface_temp_c": 17.5},
            "fields": _valid_fields(detail, "primer_coat", phase="start"),
            "photos": ["data:image/svg+xml;utf8,before-photo"],
        }
        r = requests.post(f"{API}/work-orders/{wo_id}/stages/primer_coat/start",
                          json=start_body, headers=auth, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json()["started_at"]

        # starting twice must conflict
        r2 = requests.post(f"{API}/work-orders/{wo_id}/stages/primer_coat/start",
                           json=start_body, headers=auth, timeout=15)
        assert r2.status_code == 409

        detail = _detail(auth, wo_id)
        stg = next(s for s in detail["stages"] if s["key"] == "primer_coat")
        assert stg["status"] == "in_progress"
        assert stg["start_readings"]["surface_temp_c"] == 17.5
        assert stg["start_fields"]["operator_name"]
        assert len(stg["start_photos"]) == 1

        submit_body = {
            "readings": {"end": {"ambient_temp_c": 22.0, "relative_humidity_pct": 45.0,
                                 "dew_point_c": 9.5, "surface_temp_c": 18.0}},
            "fields": _valid_fields(detail, "primer_coat", phase="end"),
            "notes": "TEST two-step", "photos": [], "result": "pass",
        }
        r3 = requests.post(f"{API}/work-orders/{wo_id}/stages/primer_coat/submit",
                           json=submit_body, headers=auth, timeout=15)
        assert r3.status_code == 200, r3.text
        assert r3.json()["result"] == "pass"

        detail = _detail(auth, wo_id)
        stg = next(s for s in detail["stages"] if s["key"] == "primer_coat")
        assert stg["status"] == "done"
        assert stg["submission"]["readings"]["start"]["surface_temp_c"] == 17.5

    def test_start_stage_not_in_case_404(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        r = requests.post(f"{API}/work-orders/{wo['work_order_id']}/stages/top_coat/start",
                          json={"readings": {}}, headers=auth, timeout=15)
        assert r.status_code == 404


# ---------- Reports & distribution ----------
class TestReports:
    def test_generate_xlsx_and_pdf(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        _submit(auth, wo["work_order_id"], "surface_prep", _valid_fields(detail, "surface_prep"))
        for fmt, magic in (("xlsx", b"PK"), ("pdf", b"%PDF")):
            r = requests.get(f"{API}/work-orders/{wo['work_order_id']}/report?format={fmt}",
                             headers=auth, timeout=30)
            assert r.status_code == 200, r.text
            assert r.content.startswith(magic), f"{fmt} magic bytes wrong"
            assert wo["work_order_id"] in r.headers.get("content-disposition", "")

    def test_report_includes_failure_reasons(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        _submit(auth, wo["work_order_id"], "surface_prep",
                _valid_fields(detail, "surface_prep", oil_water_test="NOT_OK"))
        r = requests.get(f"{API}/work-orders/{wo['work_order_id']}/report?format=xlsx",
                         headers=auth, timeout=30)
        assert r.status_code == 200
        # openpyxl re-read to confirm the reason text landed in the sheet
        import io
        from openpyxl import load_workbook
        ws = load_workbook(io.BytesIO(r.content)).active
        text = " ".join(str(c.value) for row in ws.iter_rows() for c in row if c.value)
        assert "Oil/Water" in text and "NOT OK" in text

    def test_recipients_crud_and_send_guard(self, auth):
        r = requests.post(f"{API}/report-recipients",
                          json={"name": "QA Lead", "email": "qa.lead@example.com"},
                          headers=auth, timeout=15)
        assert r.status_code == 201
        listed = requests.get(f"{API}/report-recipients", headers=auth, timeout=15).json()
        assert any(x["email"] == "qa.lead@example.com" for x in listed)
        # sending without GMAIL_* env must 503, not crash (test stack has none)
        r2 = requests.post(f"{API}/work-orders/WO-0000-0000/report/send",
                           json={"recipients": ["qa.lead@example.com"], "formats": ["pdf"]},
                           headers=auth, timeout=15)
        assert r2.status_code == 503

    def test_nov_generate_report(self, auth, spec):
        """NOV-template report: fill → PDF convert → store → download URL;
        recipients are remembered for autocomplete even when email can't send."""
        import server as srv
        if srv._find_soffice() is None:
            pytest.skip("LibreOffice not available for PDF conversion")
        wo = _create_wo(auth, spec, "only_primer")
        r = requests.post(f"{API}/work-orders/{wo['work_order_id']}/generate-report",
                          json={"recipients": ["nov.reports@example.com"]},
                          headers=auth, timeout=120)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["ok"] is True
        assert body["filename"].endswith(".pdf")
        # no GMAIL_* env on the test stack: generation still succeeds,
        # the email failure is reported instead of aborting
        assert body["email_sent"] is False
        assert "Email not configured" in (body["email_error"] or "")
        # the recipient email was remembered on first use
        listed = requests.get(f"{API}/report-recipients", headers=auth, timeout=15).json()
        assert any(x["email"] == "nov.reports@example.com" for x in listed)
        # stored PDF downloads via bearer header and via ?token= (browser flow)
        pdf = requests.get(f"{BASE}{body['download_url']}", headers=auth, timeout=30)
        assert pdf.status_code == 200
        assert pdf.content[:5] == b"%PDF-"
        tok = auth["Authorization"].split(" ", 1)[1]
        pdf2 = requests.get(f"{BASE}{body['download_url']}?token={tok}", timeout=30)
        assert pdf2.status_code == 200
        assert requests.get(f"{BASE}{body['download_url']}", timeout=30).status_code == 401
        xlsx = requests.get(f"{BASE}{body['xlsx_download_url']}", headers=auth, timeout=30)
        assert xlsx.status_code == 200
        assert xlsx.content[:2] == b"PK"


# ---------- Weather / dashboard / audit / history ----------
class TestWeather:
    def test_weather_keys(self, auth):
        r = requests.get(f"{API}/weather", headers=auth, timeout=15)
        assert r.status_code == 200
        body = r.json()
        for k in ("ambient_temp_c", "relative_humidity_pct", "dew_point_c"):
            assert k in body
            assert isinstance(body[k], (int, float))


class TestDashboard:
    def test_dashboard(self, auth, spec):
        _create_wo(auth, spec, "only_primer")  # guarantee at least one open WO
        r = requests.get(f"{API}/dashboard", headers=auth, timeout=15)
        assert r.status_code == 200
        body = r.json()
        assert body["system_status"] == "SYNCED"
        assert body["quota"]["target"] >= 1
        assert body["current_assignment"] is not None


class TestAuditAndHistory:
    def test_audit_log_has_entries(self, auth, spec):
        wo = _create_wo(auth, spec, "only_primer")
        detail = _detail(auth, wo["work_order_id"])
        _submit(auth, wo["work_order_id"], "primer_coat", _valid_fields(detail, "primer_coat"))
        time.sleep(0.5)
        r = requests.get(f"{API}/work-orders/{wo['work_order_id']}/audit-log", headers=auth, timeout=15)
        assert r.status_code == 200
        actions = {e["action"] for e in r.json()}
        assert "work_order_created" in actions
        assert "stage_submit" in actions

    def test_history_desc(self, auth):
        r = requests.get(f"{API}/inspections/history", headers=auth, timeout=15)
        assert r.status_code == 200
        items = r.json()
        assert len(items) >= 1
        ts = [it["timestamp"] for it in items]
        assert ts == sorted(ts, reverse=True)
